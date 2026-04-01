from __future__ import annotations

import json
import sqlite3
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine

from database.db_loader import build_database, process_and_load
from engine.dialogue_manager import FinancialAssistant
from knowledge.kb_builder import build_knowledge_index

from .config import (
    ATTACHMENT_4_PATH,
    ATTACHMENT_6_PATH,
    DB_PATH,
    RESULT_2_PATH,
    RESULT_3_PATH,
    RESULT_DIR,
)

REQUIRED_DB_TABLES = (
    "core_performance",
    "balance_sheet",
    "income_statement",
    "cash_flow",
)


def create_db_engine(db_path: str | Path = DB_PATH):
    return create_engine(f"sqlite:///{Path(db_path)}", future=True)


def _database_has_required_tables(conn: sqlite3.Connection) -> bool:
    try:
        rows = conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
    except sqlite3.DatabaseError:
        return False

    table_names = {str(row[0]) for row in rows if row and row[0]}
    if any(table_name not in table_names for table_name in REQUIRED_DB_TABLES):
        return False

    for table_name in REQUIRED_DB_TABLES:
        try:
            row_count = int(conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0])
        except sqlite3.DatabaseError:
            return False
        if row_count == 0:
            return False

    return True


def ensure_database(reset_database: bool = False) -> Path:
    db_path = Path(DB_PATH)
    needs_rebuild = reset_database or not db_path.exists()
    if not needs_rebuild and db_path.exists():
        try:
            conn = sqlite3.connect(db_path)
            try:
                needs_rebuild = not _database_has_required_tables(conn)
            finally:
                conn.close()
        except sqlite3.DatabaseError:
            needs_rebuild = True
    if needs_rebuild:
        build_database(db_path=db_path, reset_database=True)
    return db_path


def build_knowledge(force_rebuild: bool = False) -> None:
    build_knowledge_index(force_rebuild=force_rebuild)


def _load_question_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    parsed_rows = []
    for row in rows[1:]:
        record = {
            headers[index]: str(value).strip() if value is not None else ""
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }
        if record.get("编号"):
            parsed_rows.append(record)
    return parsed_rows


def _write_output_excel(rows: list[dict[str, Any]], output_path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    headers = ["编号", "问题", "SQL查询语句", "回答", "图形"]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def run_question_workbook(
    question_path: Path,
    output_path: Path,
    include_references: bool = False,
    reset_database: bool = False,
    rebuild_knowledge: bool = False,
) -> Path:
    ensure_database(reset_database=reset_database)
    if include_references:
        build_knowledge(force_rebuild=rebuild_knowledge)

    question_rows = _load_question_rows(question_path)
    engine = create_db_engine()
    output_rows: list[dict[str, Any]] = []

    for row in question_rows:
        question_id = row["编号"]
        turns = json.loads(row["问题"])
        assistant = FinancialAssistant(engine)
        answers = []
        sql_logs: list[str] = []
        image_logs: list[str] = []

        for turn_index, turn in enumerate(turns):
            answer_payload, sqls = assistant.process_query(turn["Q"], question_id, turn_index)
            sql_logs.extend(sqls)
            image_logs.extend(answer_payload.get("image", []) or [])

            answer_entry = {
                "Q": turn["Q"],
                "A": {
                    "content": answer_payload.get("content", ""),
                },
            }
            if answer_payload.get("image"):
                answer_entry["A"]["image"] = answer_payload["image"]
            if include_references:
                answer_entry["A"]["references"] = answer_payload.get("references", [])
            answers.append(answer_entry)

        output_rows.append(
            {
                "编号": question_id,
                "问题": row["问题"],
                "SQL查询语句": "\n\n".join(sql_logs) if sql_logs else "无",
                "回答": json.dumps(answers, ensure_ascii=False),
                "图形": json.dumps(image_logs, ensure_ascii=False) if image_logs else "无",
            }
        )

    return _write_output_excel(output_rows, output_path)


def run_task2(reset_database: bool = False) -> Path:
    return run_question_workbook(
        question_path=ATTACHMENT_4_PATH,
        output_path=RESULT_2_PATH,
        include_references=False,
        reset_database=reset_database,
    )


def run_task3(reset_database: bool = False, rebuild_knowledge: bool = True) -> Path:
    return run_question_workbook(
        question_path=ATTACHMENT_6_PATH,
        output_path=RESULT_3_PATH,
        include_references=True,
        reset_database=reset_database,
        rebuild_knowledge=rebuild_knowledge,
    )


def answer_single_question(question: str, question_id: str = "MANUAL") -> dict[str, Any]:
    ensure_database(reset_database=False)
    engine = create_db_engine()
    assistant = FinancialAssistant(engine)
    answer, sqls = assistant.process_query(question, question_id, 0)
    return {"answer": answer, "sql": sqls}


def process_single_pdf(pdf_path: str | Path, reset_database: bool = False) -> bool:
    return process_and_load(pdf_path, db_path=DB_PATH, reset_database=reset_database)


def run_task2_tests() -> Path:
    return run_task2(reset_database=False)


def run_task3_tests() -> Path:
    return run_task3(reset_database=False, rebuild_knowledge=False)

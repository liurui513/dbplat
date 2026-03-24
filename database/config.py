from __future__ import annotations

from functools import lru_cache
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data" / "data"
DB_PATH = BASE_DIR / "finance_database.db"
SCHEMA_PATH = BASE_DIR / "database" / "schema.sql"


def _first_match(pattern: str) -> Path:
    matches = sorted(DATA_DIR.glob(pattern))
    if not matches:
        raise FileNotFoundError(f"未找到匹配文件: {pattern}")
    return matches[0]


ATTACHMENT_1_PATH = _first_match("附件1：中药上市公司基本信息*.xlsx")
ATTACHMENT_3_PATH = _first_match("附件3：数据库-表名及字段说明*.xlsx")
REPORT_ROOT = _first_match("附件2*")
REPORT_PATHS = tuple(sorted(path for path in REPORT_ROOT.iterdir() if path.is_dir()))


SHEET_TO_TABLE = {
    "核心业绩指标表": "core_performance",
    "资产负债表": "balance_sheet",
    "现金流量表": "cash_flow",
    "利润表": "income_statement",
}

TABLE_ALIASES = {
    "core_performance": "core_performance_indicators_sheet",
    "balance_sheet": "balance_sheet",
    "cash_flow": "cash_flow_sheet",
    "income_statement": "income_sheet",
}

PERIOD_KEYWORDS = {
    "Q1": ("第一季度报告", "一季度报告"),
    "HY": ("半年度报告",),
    "Q3": ("第三季度报告", "三季度报告"),
    "FY": ("年度报告", "年报"),
}

PERIOD_ORDER = {
    "Q1": 1,
    "HY": 2,
    "Q3": 3,
    "FY": 4,
}

COMMON_COLUMNS = ["stock_code", "stock_abbr", "report_period", "report_year"]

FIELD_PRECISION = {
    "eps": 4,
    "net_asset_per_share": 4,
    "operating_cf_per_share": 4,
    "roe": 4,
    "roe_weighted_excl_non_recurring": 4,
    "operating_revenue_yoy_growth": 4,
    "operating_revenue_qoq_growth": 4,
    "net_profit_yoy_growth": 4,
    "net_profit_qoq_growth": 4,
    "net_profit_excl_non_recurring_yoy": 4,
    "gross_profit_margin": 4,
    "net_profit_margin": 4,
    "asset_total_assets_yoy_growth": 4,
    "liability_total_liabilities_yoy_growth": 4,
    "asset_liability_ratio": 4,
    "net_cash_flow_yoy_growth": 4,
    "operating_cf_ratio_of_net_cf": 4,
    "investing_cf_ratio_of_net_cf": 4,
    "financing_cf_ratio_of_net_cf": 4,
}


@lru_cache(maxsize=1)
def load_company_master() -> Dict[str, Dict[str, str]]:
    workbook = load_workbook(ATTACHMENT_1_PATH, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    by_code: Dict[str, Dict[str, str]] = {}
    by_abbr: Dict[str, Dict[str, str]] = {}
    by_name: Dict[str, Dict[str, str]] = {}

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row or row[1] is None:
            continue
        stock_code = str(row[1]).strip().split(".")[0].zfill(6)
        stock_abbr = (row[2] or "").strip()
        company_name = (row[3] or "").strip()
        record = {
            "stock_code": stock_code,
            "stock_abbr": stock_abbr,
            "company_name": company_name,
            "exchange": (row[6] or "").strip(),
        }
        by_code[stock_code] = record
        if stock_abbr:
            by_abbr[stock_abbr] = record
        if company_name:
            by_name[company_name] = record

    return {
        "by_code": by_code,
        "by_abbr": by_abbr,
        "by_name": by_name,
    }


@lru_cache(maxsize=1)
def load_table_columns() -> Dict[str, List[str]]:
    workbook = load_workbook(ATTACHMENT_3_PATH, read_only=True, data_only=True)
    table_columns: Dict[str, List[str]] = {}

    for sheet_name, table_name in SHEET_TO_TABLE.items():
        worksheet = workbook[sheet_name]
        columns: List[str] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            field_name = row[0]
            if field_name:
                columns.append(str(field_name).strip())
        table_columns[table_name] = columns

    return table_columns

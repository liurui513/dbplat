from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import load_workbook

from database.ocr_backend import get_ocr_backend, get_ocr_settings
from financial_assistant.config import (
    INDUSTRY_RESEARCH_DIR,
    INDUSTRY_RESEARCH_INFO_PATH,
    KNOWLEDGE_INDEX_PATH,
    PROJECT_ROOT,
    STOCK_RESEARCH_DIR,
    STOCK_RESEARCH_INFO_PATH,
)


def _canonical_name(path: Path) -> str:
    return re.sub(r"\(\d+\)|\s*-\s*副本", "", path.name).strip()


def _deduped_pdfs(root: Path) -> list[Path]:
    deduped: dict[str, Path] = {}
    for path in sorted(root.glob("*.pdf")):
        canonical = _canonical_name(path)
        current = deduped.get(canonical)
        if current is None or "(1)" in current.name:
            deduped[canonical] = path
    return sorted(deduped.values(), key=lambda item: item.name)


def _load_metadata(path: Path, key_field: str) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    metadata: dict[str, dict[str, Any]] = {}
    for row in rows[1:]:
        record = {
            headers[index]: value
            for index, value in enumerate(row)
            if index < len(headers) and headers[index]
        }
        key_value = str(record.get(key_field, "")).strip()
        if key_value:
            metadata[key_value] = record
    return metadata


def _extract_chart_caption(page_text: str) -> str | None:
    match = re.search(r"(图表\d+[:：][^\n]+)", page_text)
    return match.group(1).strip() if match else None


def _chunk_page_text(page_text: str) -> list[str]:
    normalized = re.sub(r"\n{2,}", "\n", page_text).strip()
    if not normalized:
        return []
    if "图表8：2025年国家医保目录新增7个中药产品" in normalized:
        return [normalized]
    chunks: list[str] = []
    current = []
    current_length = 0
    for line in normalized.splitlines():
        text = line.strip()
        if not text:
            continue
        if current_length + len(text) > 420 and current:
            chunks.append("\n".join(current))
            current = [text]
            current_length = len(text)
        else:
            current.append(text)
            current_length += len(text)
    if current:
        chunks.append("\n".join(current))
    return chunks


def _relative_source_path(path: Path) -> str:
    data_root = PROJECT_ROOT / "data" / "data"
    return "./" + path.relative_to(data_root).as_posix()


def _page_text_with_fallback(pdf_path: Path, pdf: pdfplumber.PDF, page_index: int) -> str:
    page = pdf.pages[page_index]
    text = (page.extract_text() or "").strip()
    ocr_backend = get_ocr_backend()
    ocr_settings = get_ocr_settings()
    if ocr_backend is None:
        return text

    if ocr_settings.fallback_only and len(text) >= ocr_settings.min_text_length:
        return text

    try:
        ocr_result = ocr_backend.extract_page(pdf_path, page_index)
    except Exception:
        return text
    return (ocr_result.text or text).strip()


def _build_documents(root: Path, source_type: str, metadata_by_title: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    documents: list[dict[str, Any]] = []
    for pdf_path in _deduped_pdfs(root):
        title = _canonical_name(pdf_path).removesuffix(".pdf")
        metadata = metadata_by_title.get(title, {})
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page_number, page in enumerate(pdf.pages, start=1):
                page_text = _page_text_with_fallback(pdf_path, pdf, page_number - 1)
                if not page_text:
                    continue
                chart_caption = _extract_chart_caption(page_text)
                for chunk_index, chunk in enumerate(_chunk_page_text(page_text), start=1):
                    documents.append(
                        {
                            "doc_id": f"{source_type}:{title}:{page_number}:{chunk_index}",
                            "source_type": source_type,
                            "title": title,
                            "file_path": str(pdf_path),
                            "relative_path": _relative_source_path(pdf_path),
                            "page_number": page_number,
                            "chunk_index": chunk_index,
                            "text": chunk,
                            "chart_caption": chart_caption,
                            "metadata": metadata,
                        }
                    )
    return documents


def build_knowledge_index(force_rebuild: bool = False) -> list[dict[str, Any]]:
    if KNOWLEDGE_INDEX_PATH.exists() and not force_rebuild:
        return load_knowledge_index()

    stock_metadata = _load_metadata(STOCK_RESEARCH_INFO_PATH, "title")
    industry_metadata = _load_metadata(INDUSTRY_RESEARCH_INFO_PATH, "title")

    documents = []
    documents.extend(_build_documents(STOCK_RESEARCH_DIR, "stock_report", stock_metadata))
    documents.extend(_build_documents(INDUSTRY_RESEARCH_DIR, "industry_report", industry_metadata))

    KNOWLEDGE_INDEX_PATH.write_text(
        json.dumps(documents, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return documents


def load_knowledge_index() -> list[dict[str, Any]]:
    if not KNOWLEDGE_INDEX_PATH.exists():
        return build_knowledge_index(force_rebuild=True)
    return json.loads(KNOWLEDGE_INDEX_PATH.read_text(encoding="utf-8"))
